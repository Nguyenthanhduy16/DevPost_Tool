"""Chạy script này để tạo file Excel mẫu: python scripts/create_sample.py"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = [
    "Job Title", "Job Category", "Skills", "Job Type", "Role",
    "Level", "Contract Type", "Tỉnh", "Job Detail",
    "Mô tả CV", "Yêu cầu công việc", "LƯƠNG MIN", "LƯƠNG MAX",
]

SAMPLE = [
    ["Chuyên viên Khách hàng Cá nhân - Long Biên HN (2026TD451031)",
     "Banking", "Customer Service, Communication, Sales",
     "Full-time", "Individual Contributor", "Junior", "Permanent", "Hà Nội",
     "Tham gia trực tiếp vào các hoạt động kinh doanh, tư vấn và chăm sóc khách hàng cá nhân tại khu vực Long Biên.",
     "Tư vấn và bán các sản phẩm tài chính cá nhân\nChăm sóc và duy trì quan hệ khách hàng\nThực hiện các chỉ tiêu kinh doanh được giao",
     "Tốt nghiệp Đại học chuyên ngành Tài chính, Ngân hàng hoặc liên quan\nKỹ năng giao tiếp và thuyết phục tốt\nƯu tiên có kinh nghiệm trong lĩnh vực tài chính ngân hàng",
     "9000000", "25000000"],

    ["Senior Data Engineer",
     "Information Technology", "Python, Spark, SQL, Kafka, Airflow",
     "Full-time", "Senior", "Professional", "Permanent", "Hồ Chí Minh",
     "Xây dựng và vận hành hệ thống data pipeline cho nền tảng tài chính quy mô lớn.",
     "Thiết kế và triển khai data pipeline với Spark và Kafka\nTối ưu hóa hiệu suất truy vấn SQL\nXây dựng Data Warehouse và Data Lake",
     "Tối thiểu 3 năm kinh nghiệm Data Engineering\nThành thạo Python, PySpark, SQL\nKinh nghiệm với cloud AWS/GCP/Azure",
     "15000000", "35000000"],
]

def create():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JD Data"

    hfill = PatternFill("solid", fgColor="4F46E5")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin  = Side(style="thin", color="DDDDDD")
    bd    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bd

    for r, row in enumerate(SAMPLE, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = bd

    col_widths = [55,20,30,14,20,14,16,14,40,40,40,14,14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    out = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "jd_data.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"✅ Tạo file mẫu: {out}")
    print(f"   {len(SAMPLE)} dòng dữ liệu mẫu")
    print(f"\nCác cột cần điền:")
    for h in HEADERS:
        print(f"  • {h}")

if __name__ == "__main__":
    create()
